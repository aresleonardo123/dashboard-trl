# funciones.py
import pandas as pd
import io
import re
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generar_excel_aprobados(df: pd.DataFrame) -> io.BytesIO:
    columnas = [
        "Nombre del Proyecto", "Aprobado", "Nivel TRL", "Segmento TRL", "Docente Acompañante",
        "Ubicación", "Nivel de Inglés", "Puntaje TRL 1-3", "Puntaje TRL 4-7", "Puntaje TRL 8-9",
        "Puntaje Total", "Insights"
    ]

    # Asegurar que todas las columnas existan
    for col in columnas:
        if col not in df.columns:
            df[col] = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos Aprobados"

    # Estilos
    header_fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Escribir encabezados
    for col_num, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_style

    # Escribir datos
    for row_num, row_data in enumerate(df[columnas].values.tolist(), 2):
        for col_num, value in enumerate(row_data, 1):
            # Convertir listas a string plano
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, (dict, tuple)):
                value = str(value)

            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=False)
            cell.border = border_style

    # (Opcional) Ancho fijo por columna
    for col_letter in ws.iter_cols(min_row=1, max_row=1):
        letra = col_letter[0].column_letter
        ws.column_dimensions[letra].width = 25  # puedes ajustar a tu gusto

    return _guardar_excel(wb)

def _guardar_excel(workbook: Workbook) -> io.BytesIO:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def normalizar_contraseña(password: str) -> str:
    """
    Normaliza la contraseña reemplazando todos los espacios y caracteres especiales por guiones
    y asegura el formato consistente.
    """
    if not password:
        return ""
    
    # Paso 1: Reemplazar cualquier tipo de espacio o carácter especial por guión
    password = re.sub(r'[\s\u00A0\u200B\u200C\u200D\uFEFF\W]+', '-', password)
    
    # Paso 2: Eliminar guiones consecutivos y al inicio/final
    password = re.sub(r'-+', '-', password).strip('-')
    
    # Paso 3: Asegurar formato de 6 partes con guiones
    parts = password.split('-')
    if len(parts) >= 6:
        return '-'.join(parts[:6])  # Toma los primeros 6 segmentos
    return password

def segmento_trl(nivel):
    if 1 <= nivel <= 3:
        return "TRL 1-3"
    elif 4 <= nivel <= 7:
        return "TRL 4-7"
    elif 8 <= nivel <= 9:
        return "TRL 8-9"
    return "Desconocido"

def calcular_puntajes_por_segmento(fila, diccionario):
    puntajes = {"TRL 1-3": 0, "TRL 4-7": 0, "TRL 8-9": 0}
    for pregunta, respuestas_map in diccionario.items():
        for columna, respuesta_usuario in fila.items():
            if respuesta_usuario in respuestas_map:
                datos = respuestas_map[respuesta_usuario]
                puntajes[datos["segmento"]] += datos["puntaje"]
    return puntajes

def analizar_madurez(proyecto, segmento):
    """Analiza el nivel de madurez del proyecto basado en múltiples factores."""
    trl = proyecto["Nivel TRL"]
    docente = proyecto["Docente Acompañante"]
    puntajes = {
        "TRL 1-3": proyecto["Puntaje TRL 1-3"],
        "TRL 4-7": proyecto["Puntaje TRL 4-7"],
        "TRL 8-9": proyecto["Puntaje TRL 8-9"]
    }
    
    if segmento == "TRL 1-3":
        if puntajes["TRL 1-3"] >= 40:
            return "Investigación sólida: Buen fundamento teórico y validación inicial"
        else:
            return "Etapa conceptual: Necesita más desarrollo teórico y validación"
    
    elif segmento == "TRL 4-7":
        if puntajes["TRL 4-7"] >= 50:
            return "Prototipo funcional: Validación técnica en progreso"
        else:
            return "Prototipo inicial: Requiere más desarrollo técnico"
    
    else:  # TRL 8-9
        if puntajes["TRL 8-9"] >= 60:
            return "Listo para implementación: Alta preparación para el mercado"
        else:
            return "Casi listo: Necesita ajustes finales para implementación"

def identificar_fortalezas(proyecto):
    """Identifica las principales fortalezas del proyecto."""
    fortalezas = []
    
    # Fortalezas por puntajes
    if proyecto["Puntaje TRL 1-3"] >= 40:
        fortalezas.append("✅ Innovación bien fundamentada con investigación sólida")
    if proyecto["Puntaje TRL 4-7"] >= 50:
        fortalezas.append("✅ Desarrollo técnico avanzado y validado")
    if proyecto["Puntaje TRL 8-9"] >= 50:
        fortalezas.append("✅ Alto potencial de implementación y escalabilidad")
    
    # Otras fortalezas
    if proyecto["Docente Acompañante"]:
        fortalezas.append("✅ Excelente acompañamiento académico")
    if proyecto.get("Nivel de Inglés", "") in ["Avanzado", "Intermedio"]:
        fortalezas.append("✅ Buena capacidad para documentación internacional")
    
    return fortalezas

def identificar_debilidades(proyecto):
    """Identifica las principales debilidades del proyecto."""
    debilidades = []
    
    # Debilidades por puntajes
    if proyecto["Puntaje TRL 1-3"] < 30:
        debilidades.append("⚠️ Fundamentación teórica débil - necesita más investigación")
    if proyecto["Puntaje TRL 4-7"] < 40:
        debilidades.append("⚠️ Desarrollo técnico insuficiente - requiere más validación")
    if proyecto["Puntaje TRL 8-9"] < 40:
        debilidades.append("⚠️ Preparación para el mercado limitada - necesita más desarrollo")
    
    # Otras debilidades
    if not proyecto["Docente Acompañante"]:
        debilidades.append("⚠️ Falta acompañamiento docente - recomendar mentoría")
    if proyecto.get("Nivel de Inglés", "") == "Básico":
        debilidades.append("⚠️ Limitaciones en inglés - afecta potencial internacional")
    
    return debilidades

def generar_recomendaciones(proyecto):
    """Genera recomendaciones personalizadas para el proyecto."""
    recomendaciones = []
    segmento = segmento_trl(proyecto["Nivel TRL"])
    
    # Recomendaciones por segmento TRL
    if segmento == "TRL 1-3":
        recomendaciones.append("Priorizar investigación y validación conceptual")
        if proyecto["Puntaje TRL 1-3"] < 30:
            recomendaciones.append("Realizar más investigación de mercado y técnica")
    elif segmento == "TRL 4-7":
        recomendaciones.append("Enfocarse en desarrollo técnico y pruebas")
        if proyecto["Puntaje TRL 4-7"] < 40:
            recomendaciones.append("Realizar pruebas técnicas más rigurosas")
    else:
        recomendaciones.append("Preparar estrategia de implementación y comercialización")
        if proyecto["Puntaje TRL 8-9"] < 50:
            recomendaciones.append("Realizar pruebas piloto con usuarios finales")
    
    # Recomendaciones específicas
    if not proyecto["Docente Acompañante"]:
        recomendaciones.append("Buscar mentoría docente para fortalecer el proyecto")
    if proyecto.get("Nivel de Inglés", "") == "Básico":
        recomendaciones.append("Mejorar documentación en inglés para mayor impacto")
    
    return recomendaciones

def evaluar_potencial(proyecto):
    """Evalúa el potencial general del proyecto."""
    puntaje_total = sum([
        proyecto["Puntaje TRL 1-3"],
        proyecto["Puntaje TRL 4-7"],
        proyecto["Puntaje TRL 8-9"]
    ])
    
    if puntaje_total >= 120:
        return "🌟 Excelente potencial: Proyecto bien desarrollado en todas las áreas"
    elif puntaje_total >= 80:
        return "✨ Buen potencial: Proyecto sólido con algunas áreas para mejorar"
    elif puntaje_total >= 50:
        return "💡 Potencial moderado: Necesita trabajo en varias áreas"
    else:
        return "🔍 Potencial limitado: Requiere desarrollo significativo"

 # Modificar la función generar_insights para usar todas las columnas necesarias

def generar_insights(proyecto, diccionario=None):
    """Genera insights personalizados (versión unificada)"""
    insights = []
    
    # Asegurar existencia de campos
    proyecto["Industria"] = proyecto.get("Industria", "No especificada")
    proyecto["Ubicación"] = proyecto.get("Ubicación", "No especificada")
    
    # Resto de la lógica igual que en main.py
    segmento = segmento_trl(proyecto["Nivel TRL"])
    insights.append(analizar_madurez(proyecto, segmento))
    insights.extend(identificar_fortalezas(proyecto))
    insights.extend(identificar_debilidades(proyecto))
    insights.extend(generar_recomendaciones(proyecto))
    insights.append(evaluar_potencial(proyecto))
    
    # Insight adicional de industria
    insights.append(f"🏭 Sector: {proyecto['Industria']} - Considerar tendencias del mercado relacionadas")
    
    return insights
